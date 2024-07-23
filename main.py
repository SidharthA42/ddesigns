from kivy.lang import Builder
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.label import MDLabel
from kivymd.uix.textfield import MDTextField
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.gridlayout import MDGridLayout
from datetime import datetime
import pandas as pd
import os
import platform
import subprocess
from openpyxl import load_workbook

KV = '''
BoxLayout:
    orientation: 'vertical'
    padding: 20
    spacing: 20

    MDBoxLayout:
        orientation: 'vertical'
        adaptive_height: True
        spacing: 10

        MDLabel:
            text: "Site:"
            theme_text_color: "Primary"
            halign: "center"

        MDTextField:
            id: site_input
            hint_text: "Enter site name"
            mode: "rectangle"
            size_hint_x: 0.9
            pos_hint: {"center_x": 0.5}

        MDLabel:
            text: "Date (dd/mm/yyyy):"
            theme_text_color: "Primary"
            halign: "center"

        MDTextField:
            id: date_input
            hint_text: "Enter date"
            mode: "rectangle"
            size_hint_x: 0.9
            pos_hint: {"center_x": 0.5}

        MDLabel:
            text: "Debit:"
            theme_text_color: "Primary"
            halign: "center"

        MDTextField:
            id: debit_input
            hint_text: "Enter debit amount"
            mode: "rectangle"
            size_hint_x: 0.9
            pos_hint: {"center_x": 0.5}

        MDLabel:
            text: "Credit:"
            theme_text_color: "Primary"
            halign: "center"

        MDTextField:
            id: credit_input
            hint_text: "Enter credit amount"
            mode: "rectangle"
            size_hint_x: 0.9
            pos_hint: {"center_x": 0.5}

        MDLabel:
            text: "Description:"
            theme_text_color: "Primary"
            halign: "center"

        MDTextField:
            id: description_input
            hint_text: "Enter description"
            mode: "rectangle"
            size_hint_x: 0.9
            pos_hint: {"center_x": 0.5}

    MDBoxLayout:
        orientation: 'horizontal'
        size_hint_y: None
        height: self.minimum_height
        spacing: 10
        padding: [20, 0, 20, 0]

        MDRaisedButton:
            text: "Enter"
            on_press: app.save_data()

        MDRaisedButton:
            text: "View Data"
            on_press: app.view_data()

        MDRaisedButton:
            text: "Download Data"
            on_press: app.download_data()

        MDRaisedButton:
            text: "Clear Data"
            on_press: app.clear_data()

    MDScrollView:
        size_hint: (1, 1)
        MDGridLayout:
            id: output_layout
            cols: 1
            adaptive_height: True
'''

class DDesignsApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "BlueGray"
        return Builder.load_string(KV)

    def save_data(self):
        site = self.root.ids.site_input.text
        date = datetime.strptime(self.root.ids.date_input.text, '%d/%m/%Y')
        debit = float(self.root.ids.debit_input.text)
        credit = float(self.root.ids.credit_input.text)
        description = self.root.ids.description_input.text

        data = {
            'Date': [date],
            'Debit': [debit],
            'Credit': [credit],
            'Description': [description]
        }

        df = pd.DataFrame(data)

        filename = f'{site}.xlsx'
        if os.path.exists(filename):
            existing_df = pd.read_excel(filename)
            df = pd.concat([existing_df, df], ignore_index=True)

        df.to_excel(filename, index=False)

        # Fixing column width and date format
        book = load_workbook(filename)
        sheet = book.active
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.coordinate in sheet.merged_cells:  # not check merge_cells
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        for cell in sheet['A']:  # Assuming 'A' column is for dates
            cell.number_format = 'DD/MM/YYYY'

        book.save(filename)

        self.root.ids.site_input.text = ''
        self.root.ids.date_input.text = ''
        self.root.ids.debit_input.text = ''
        self.root.ids.credit_input.text = ''
        self.root.ids.description_input.text = ''

    def view_data(self):
        self.root.ids.output_layout.clear_widgets()
        site = self.root.ids.site_input.text
        filename = f'{site}.xlsx'

        if os.path.exists(filename):
            df = pd.read_excel(filename)

            for index, row in df.iterrows():
                row_str = f"Date: {row['Date']} | Debit: {row['Debit']} | Credit: {row['Credit']} | Description: {row['Description']}"
                self.root.ids.output_layout.add_widget(MDLabel(text=row_str, size_hint_y=None, height=40))
        else:
            self.root.ids.output_layout.add_widget(MDLabel(text="No data found for this site.", size_hint_y=None, height=40))

    def download_data(self):
        site = self.root.ids.site_input.text
        filename = f'{site}.xlsx'

        if os.path.exists(filename):
            if platform.system() == "Windows":
                os.startfile(os.path.abspath(filename))
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", os.path.abspath(filename)])
            else:
                subprocess.Popen(["xdg-open", os.path.abspath(filename)])
        else:
            self.root.ids.output_layout.clear_widgets()
            self.root.ids.output_layout.add_widget(MDLabel(text="No data found for this site to download.", size_hint_y=None, height=40))

    def clear_data(self):
        site = self.root.ids.site_input.text
        filename = f'{site}.xlsx'

        if os.path.exists(filename):
            os.remove(filename)
            self.root.ids.output_layout.clear_widgets()
            self.root.ids.output_layout.add_widget(MDLabel(text="Data cleared for this site.", size_hint_y=None, height=40))
        else:
            self.root.ids.output_layout.clear_widgets()
            self.root.ids.output_layout.add_widget(MDLabel(text="No data found for this site to clear.", size_hint_y=None, height=40))


if __name__ == '__main__':
    DDesignsApp().run()
